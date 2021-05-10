# CALCULOS PRINCIPALES
## CÁLCULO de DnT y R' - ISO 140-4

from openpyxl import load_workbook      # Nos permite leer de ficheros .xlsx
import math                             # Nos permite usar funciones matemáticas

# VARIABLES GLOBALES
SHEET = 'Hoja 1'        # Hoja del archivo XLSX
FILE  = 'datos2.xlsx'    # Variable para el fichero principal
FILE_TR = 'TR.xlsx'     # Variable para un segundo fichero (TR)
T = 0.5                 # Tiempo de referencia = 0.5s
V = 46.3                # Volumen del recinto receptor
S = 15.4                # Área del elemento separador


## ARRAYS PARA TRATAR DATOS DE MANERA MÁS SENCILLA:
# Rango de frecuencias de interes:
arrayFR = [50, 63, 80, 100, 125, 160, 200, 250, 315, 400, 500, 630, 800, 1000, 1250, 1600, 2000, 2500, 3150, 4000, 5000]

# Arrays vacíos
LpE = []           # Array para los resultados del nivel de presión en emisión, LpE
LpR = []           # Array para los resultados del nivel de presión en recepción, LpR
TR = []           # Array para los resultados del tiempo de reverberación en recepción, TR con paneles
myDnT = []           # Array para los resultados del nivel de diferencia normalizada, DnT
myR = []            # Array para los resultados del índice de reducción sonora aparente, R'


# FÓRMULAS:
# Cálculo del promedio de posiciones de micrófono para cada banda de frecuencia, Lp, donde
# A: min_row; B: max_row; C: min_col; D: max_col; n: posiciones de micrófono
def toCalc(A, B, C, D, n, myArray):
    wb = load_workbook(FILE)        # Cargamos en wb el fichero
    sheet = wb[SHEET]               # Cargamos la hoja del fichero de donde obtenemos los datos

    for value in sheet.iter_rows(min_row =A, max_row =B, min_col = C, max_col =D,
                                 values_only=True):
        i = 0
        L = 0
        for cell in value:
            L = L + 10**(value[i]/10)       # calculo 10*(Lp/10) y sumo al resultado anterior.
            i = i + 1
        L = 10*math.log(L/n,(10))      # Calculo el promedio logaritmico de la suma total, L.
        myArray.append(round(L,1))          # Array ordenado de datos


# Cálculo del tiempo de reverberación TR20 para cada banda de frecuencia:
def toRT(A, B, C, D, myArray):
    wb = load_workbook(FILE_TR)        # Cargamos en wb el fichero
    sheet = wb[SHEET]                  # Cargamos la hoja del fichero de donde obtenemos los datos

    for value in sheet.iter_rows(min_row =A, max_row =B, min_col = C, max_col =D,
                                 values_only=True):
        RT = 0
        i = 0
        n = 0           #NOTA: Se añade un contadores de valores porque hay algunos resultados nulos, por lo que no todas las
                        # medidas tienen el mismo número de promedios.
        for cell in value:
            if value != '*':            # Cuando los valores  no son nulos:
                RT = RT + value[i]      # calculo el promedio de los valores del TR por cada posición
                n = n + 1               # Añado 1 a mis valores 'no nulos'
            i = i + 1
        RT = RT/n                       # TR romedio para la frecuencia x
        myArray.append(round(RT,1))      # Array ordenado de datos


# Cálculo de la diferencia estandarizada para cada banda de frecuencia, DnT:
# donde A representa el Lp en el recinto EMISOR y B el Lp en el recinto RECEPTOR.
def toDnT (A, B, TR, myArray):
    i = 0
    for num in A:
        DnT = A[i] - B[i] + 10*math.log(TR[i]/T, (10))
        i = i + 1
        myArray.append(round(DnT,1))


# Cálculo de la reducción sonora aparente para cada banda de frecuencia, R':
# donde A representa el Lp en el recinto EMISOR y B el Lp en el recinto RECEPTOR.
def toR(A, B, TR, myArray):
    i = 0
    for num in A:
        Abs = (0.161*V)/TR[i]                         # Calculo del área de absorción equivalente
        R = A[i] - B[i] + 10*math.log(S/Abs,(10))
        i = i + 1
        myArray.append(round(R,1))

# Imprime resultados por pantalla:
def toPrint(myArray, legend):
    i = 0                                             # Iterador para los arrays
    for num in arrayFR:
        print(myArray[i], legend , arrayFR[i], 'Hz')  # Se imprime por terminal el elemento i del array
        i = i + 1                                     # Siguiente elemento
                               # Siguiente elemento

if __name__ == "__main__":
    print()
    print('NIVELES EN EMISIÓN')
    print('------------------')
    toCalc(8, 28, 3, 12, 10, LpE)
    toPrint(LpE, 'dB - ')

    print()
    print('NIVELES EN RECEPCIÓN')
    print('--------------------')
    toCalc(8, 28, 16, 25, 10, LpR)
    toPrint(LpR, 'dB - ')

    print()
    print('TR | PANELES ABSORBENTES')
    print('------------------------')
    toRT(6, 26, 3, 8, TR)
    toPrint(TR, 's - ')

    print()
    print('DIFERENCIA ESTANDARIZADA')
    print('------------------------')
    toDnT(LpE, LpR, TR, myDnT)
    toPrint(myDnT, 'dB - ')

    print()
    print('ÍNDICE DE REDUCCIÓN SONORA APARENTE')
    print('-----------------------------------')
    toR(LpE, LpR, TR, myR)
    toPrint(myR, 'dB - ')
