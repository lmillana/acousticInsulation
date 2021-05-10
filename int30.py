# CALCULOS PRINCIPALES
## CÁLCULO de DnT y R' POR INTENSIMETRÍA - ISO 15186-2
## 30CM ENTRE SONDA Y PARED DE SEPARACIÓN - EXPERIMENTO 1


from openpyxl import load_workbook      # Nos permite leer de ficheros .xlsx
import math                             # Nos permite usar funciones matemáticas
from pylab import *                     # Nos permite crear gráficos

# Variables globales
SHEET = 'Hoja 1'             # Hoja del archivo XLSX
FILE = 'datos2.xlsx'         # Variable para el fichero principal
S = 15.4                     # Área total del elemento de separación sometido a ensayo
So = 1                       # [m2]
Sm = 9.2                     # Área total de la superfecie de medición [m2]
Smi = 4.6                    # Área respectiva de las sub-áreas 1 y 2 [m2]
Smj = 1.2                    # Área de las sub-áreas laterales [m2]
Ao = 10                      # [m2]
c = 343
N = 2                        # Número de posiciones de altavoz
Io = 10**(-12)               # Intensidad acústica de referencia [W/m2]
Sb = 81.9                    # Área de todas las superficies límite en el recinto receptor
V = 46.3                     # Volumen del recinto receptor

## ARRAYS PARA TRATAR DATOS DE MANERA MÁS SENCILLA:
# Rango de frecuencias de interes:
FR = [50, 63, 80, 100, 125, 160, 200, 250, 315, 400, 500, 630, 800, 1000, 1250, 1600, 2000, 2500, 3150, 4000, 5000]
# Valores en Emisión - SB:
E = [69.1, 67.5, 67.4, 85.3, 85.7, 92.7, 96.5, 98.3, 100.4, 101.1, 100.3, 100.4, 99.6, 99.0, 98.8, 98.1, 99.1, 101.4, 99.7, 96.7, 95.5]


## SUB-ÁREAS DEL ELEMENTO DE MEDICIÓN:
SUB11 = [] # Sub-área 1, posición de altavoz A1
SUB12 = [] # Sub-área 1, posición de altavoz A2
SUB21 = [] # Sub-área 2, posición de altavoz A1
SUB22 = [] # Sub-área 2, posición de altavoz A2

## SUB-ÁREAS LATERALES:
arrayS3 = [] # Array para los resultados de la sub-área 3
arrayS4 = [] # Array para los resultados de la sub-área 4
arrayS5 = [] # Array para los resultados de la sub-área 5
arrayS6 = [] # Array para los resultados de la sub-área 6

LIn = [] # Array para los resultados combinados de ambas sub-áreas
DIn = [] # Array para los resultados de la diferencia normalizada por intensimetría
RI = []  # Array para los resultados de R'I SIN influencia de paredes laterales
RI2 = []  # Array para los resultados de R'I CON influencia de paredes laterales

# Imprime resultados por pantalla:
def toPrint(myArray):
    i = 0                                        # Iterador para los arrays
    for num in FR:
        print(myArray[i], 'dB - ' , FR[i], 'Hz') # Se imprime por terminal el elemento i del array
        i = i + 1                                # Siguiente elemento

# Cálculo del promedio de posiciones de micrófono para cada banda de frecuencia, Lp, donde
# A: min_row; B: max_row; C: min_col; D: max_col; n: posiciones de micrófono
def toSUB(A, B, C, D, myArray):
    wb = load_workbook(FILE)        # Cargamos en wb el fichero
    sheet = wb[SHEET]               # Cargamos la hoja del fichero de donde obtenemos los datos

    for value in sheet.iter_rows(min_row =A, max_row =B, min_col = C, max_col = D,
                                 values_only=True):
        L = 0
        i = 0
        for cell in value:
            L = L + value[i]
            i = i + 1
        myArray.append(round(L/2,1))
    toPrint(myArray)
    print()

# Calculo de la combinación de resultados:
# A: A1, sub-área 1; B: A1, sub-área 2; C: A2, sub-área 1; D: A2, sub-área 2.
def toLIn(A, B, C, D, myArray):
    i = 0                                       # Iterador para los arrays
    for num in A:
        In = (Io/N)*((1/Sm)*(Smi*(10**(0.1*A[i])+ 10**(0.1*B[i])+ 10**(0.1*C[i])+ 10**(0.1*D[i]))))
        LIn = 10*math.log((In/Io),(10))         # Cálculo del nivel de intensidad superficial:
        i = i + 1                               # Siguiente elemento del array
        myArray.append(round(LIn,1))            # Añadimos al array
    toPrint(myArray)                            # Imprimimos resultados
    print()


# Cálculo de la diferencia de nivel estandarizada por intensimetría para cada banda de frecuencia, DIn:
# donde A: LpE; B: LIn en el recinto RECEPTOR.
def toDIn(A, B, myArray):
    i = 0                                       # Iterador para los arrays
    for num in A:
        DIn = (A[i] - 6) - (B[i] + 10*math.log((Sm/Ao),(10)))
        i = i + 1                               # Siguiente elemento del array
        myArray.append(round(DIn, 1))           # Añadimos al array
    toPrint(myArray)                            # Imprimimos resultados
    print()

# Cálculo de la reducción sonora aparente por intensimetría para cada banda de frecuencia, R'I:
# donde A: LpE; B: LIn; sin influencia por paredes laterales.
def toRI(A, B, FR, myArray):
    i = 0                                       # Iterador para los arrays
    for num in A:
        x = 1 + ((Sb*(c/FR[i]))/(8*V))
        Kc = 10*math.log(x,(10))
        sum = Sm*(10**(0.1*B[i]))
        RI = (A[i] - 6 + 10*math.log(S/So,(10))) - 10*math.log(sum,(10)) + Kc

        i = i + 1                               # Siguiente elemento del array
        myArray.append(round(RI, 1))            # Añadimos al array
    toPrint(myArray)                            # Imprimimos resultados
    print()

"""
# Cálculo de la reducción sonora aparente por intensimetría para cada banda de frecuencia, R'I:
# donde A representa el Lp en el recinto EMISOR, B el nivel de intensidad en el recinto RECEPTOR y
# A: LpE; B: LIn; C: sub3; D: sub4; E: sub5; F: sub6.
def toRI2(A, B, C, D, E, F, FR):
    landa = c/FR
    x = 1 + ((Sb*landa)/(8*V))
    Kc = 10*math.log(x,(10))
    sum = Sm*(10**(0.1*B)) + Smj*(10**(0.1*C)+ 10**(0.1*D) + 10**(0.1*E) + 10**(0.1*F))
    RI = (A - 6 + 10*math.log(S/So,(10))) - 10*math.log(sum,(10)) + Kc
    return(RI)
"""

if __name__ == "__main__":
    print('SUB-ÁREA 1 - A1')
    print('---------------')
    toSUB(42, 62, 3, 4, SUB11)

    print('SUB-ÁREA 1 - A2')
    print('---------------')
    toSUB(42, 62, 5, 6, SUB12)

    print('SUB-ÁREA 2 - A1')
    print('---------------')
    toSUB(42, 62, 7, 8, SUB21)

    print('SUB-ÁREA 2 - A2')
    print('---------------')
    toSUB(42, 62, 9, 10, SUB22)

    print()
    print('COMBINACIÓN DE RESULTADOS - LIn')
    print('-------------------------------')
    toLIn(SUB11, SUB12, SUB21, SUB22, LIn)

    print('DIFERENCIA DE NIVEL NORMALIZADO POR INTENSIMETRÍA')
    print('-------------------------------------------------')
    toDIn(E, LIn, DIn)

    print("R'I SIN INFLUENCIA DE ELEMENTOS LATERALES")
    print('------------------------------------------')
    toRI(E, LIn, FR, RI)

"""
    print("R'I CON INFLUENCIA DE ELEMENTOS LATERALES")
    print('------------------------------------------')
    toAdd('C74','C94',arrayS3)                # Creamos array con los valores SUB3
    toAdd('D74','D94', arrayS4)               # Creamos array con los valores SUB4
    toAdd('E74','E94',arrayS5)                # Creamos array con los valores SUB5
    toAdd('F74','F94', arrayS6)               # Creamos array con los valores SUB6
    toPrint3(arrayE, arrayLIn, arrayS3, arrayS4, arrayS5, arrayS6, arrayFR, toRI2, arrayRI2)
"""
